from __future__ import annotations

import sqlite3
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
import zipfile
import logging
from logging.handlers import RotatingFileHandler
import os
import secrets
import time
from dotenv import load_dotenv

from flask import Flask, jsonify, redirect, render_template, request, send_from_directory, session, url_for
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Load environment variables from .env file
load_dotenv()

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "maintenance.db"
LOG_DIR = BASE_DIR / "logs"
LOG_FILE = LOG_DIR / "app.log"

# Create logs directory if it doesn't exist
LOG_DIR.mkdir(exist_ok=True)

app = Flask(__name__)
app.secret_key = "change-this-secret-key"

# SendGrid Configuration for OTP
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY", "")
SENDGRID_FROM_EMAIL = os.getenv("SENDGRID_FROM_EMAIL", "manitenance.details@gmail.com")

# Initialize SendGrid client if API key is available
sg_client = None
if SENDGRID_API_KEY:
    try:
        sg_client = SendGridAPIClient(SENDGRID_API_KEY)
    except Exception:
        sg_client = None

# OTP Configuration
OTP_VALIDITY_MINUTES = 5
OTP_LENGTH = 6

# Configure logging
log_formatter = logging.Formatter(
    '%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# File handler with rotation (max 5MB per file, keep 10 files)
file_handler = RotatingFileHandler(
    str(LOG_FILE),
    maxBytes=5*1024*1024,  # 5MB
    backupCount=10
)
file_handler.setFormatter(log_formatter)
file_handler.setLevel(logging.INFO)

# Add handler to Flask app logger
app.logger.addHandler(file_handler)
app.logger.setLevel(logging.INFO)


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, column_def: str) -> None:
    cols = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    if column_name not in [c["name"] for c in cols]:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_def}")


def migrate_units_table_if_needed(conn: sqlite3.Connection) -> None:
    schema_row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'units'"
    ).fetchone()
    if not schema_row or not schema_row["sql"]:
        return

    schema_sql = schema_row["sql"].upper()
    if "UNIT_NUMBER TEXT NOT NULL UNIQUE" not in schema_sql:
        return

    conn.executescript(
        """
        ALTER TABLE units RENAME TO units_old;

        CREATE TABLE units (
            unit_id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_number TEXT NOT NULL,
            resident_name TEXT NOT NULL,
            tenant_name TEXT,
            monthly_maintenance REAL NOT NULL,
            apartment_id INTEGER,
            FOREIGN KEY (apartment_id) REFERENCES apartments(apartment_id)
        );

        INSERT INTO units (unit_id, unit_number, resident_name, tenant_name, monthly_maintenance, apartment_id)
        SELECT unit_id, unit_number, resident_name, NULL, monthly_maintenance, apartment_id
        FROM units_old;

        DROP TABLE units_old;
        """
    )


def init_db() -> None:
    conn = get_conn()
    cur = conn.cursor()

    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS apartments (
            apartment_id INTEGER PRIMARY KEY AUTOINCREMENT,
            apartment_name TEXT NOT NULL,
            owner_user_id INTEGER NOT NULL,
            apartment_name_locked INTEGER NOT NULL DEFAULT 0,
            opening_balance_locked INTEGER NOT NULL DEFAULT 0,
            default_maintenance_amount REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (owner_user_id) REFERENCES users(user_id)
        );

        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            display_name TEXT NOT NULL,
            role TEXT NOT NULL,
            apartment_id INTEGER,
            FOREIGN KEY (apartment_id) REFERENCES apartments(apartment_id)
        );

        CREATE TABLE IF NOT EXISTS units (
            unit_id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_number TEXT NOT NULL,
            resident_name TEXT NOT NULL,
            tenant_name TEXT,
            monthly_maintenance REAL NOT NULL,
            apartment_id INTEGER,
            FOREIGN KEY (apartment_id) REFERENCES apartments(apartment_id)
        );

        CREATE TABLE IF NOT EXISTS payments (
            payment_id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_id INTEGER NOT NULL,
            month INTEGER NOT NULL,
            year INTEGER NOT NULL,
            amount_paid REAL NOT NULL,
            payment_date TEXT NOT NULL,
            payment_mode TEXT NOT NULL,
            transaction_ref TEXT,
            apartment_id INTEGER,
            FOREIGN KEY (unit_id) REFERENCES units(unit_id),
            FOREIGN KEY (apartment_id) REFERENCES apartments(apartment_id)
        );

        CREATE TABLE IF NOT EXISTS maintenance_bills (
            bill_id INTEGER PRIMARY KEY AUTOINCREMENT,
            unit_id INTEGER NOT NULL,
            month INTEGER NOT NULL,
            year INTEGER NOT NULL,
            bill_amount REAL NOT NULL,
            apartment_id INTEGER,
            FOREIGN KEY (unit_id) REFERENCES units(unit_id),
            FOREIGN KEY (apartment_id) REFERENCES apartments(apartment_id),
            UNIQUE(unit_id, month, year, apartment_id)
        );

        CREATE TABLE IF NOT EXISTS expenditures (
            expense_id INTEGER PRIMARY KEY AUTOINCREMENT,
            expense_date TEXT NOT NULL,
            category TEXT NOT NULL,
            vendor_name TEXT NOT NULL,
            amount REAL NOT NULL,
            invoice_ref TEXT,
            remarks TEXT,
            apartment_id INTEGER,
            FOREIGN KEY (apartment_id) REFERENCES apartments(apartment_id)
        );

        CREATE TABLE IF NOT EXISTS income (
            income_id INTEGER PRIMARY KEY AUTOINCREMENT,
            income_date TEXT NOT NULL,
            category TEXT NOT NULL,
            description TEXT NOT NULL,
            amount REAL NOT NULL,
            remarks TEXT,
            apartment_id INTEGER,
            FOREIGN KEY (apartment_id) REFERENCES apartments(apartment_id)
        );

        CREATE TABLE IF NOT EXISTS app_settings (
            setting_key TEXT PRIMARY KEY,
            setting_value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS otp_sessions (
            otp_id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone_number TEXT NOT NULL,
            otp_code TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            expires_at TEXT NOT NULL,
            is_verified INTEGER NOT NULL DEFAULT 0,
            attempts INTEGER NOT NULL DEFAULT 0
        );
        """
    )

    # Migration support for older schema versions.
    ensure_column(conn, "users", "apartment_id", "INTEGER")
    ensure_column(conn, "units", "apartment_id", "INTEGER")
    ensure_column(conn, "units", "tenant_name", "TEXT")
    ensure_column(conn, "payments", "apartment_id", "INTEGER")
    ensure_column(conn, "expenditures", "apartment_id", "INTEGER")
    ensure_column(conn, "income", "apartment_id", "INTEGER")
    ensure_column(conn, "apartments", "apartment_name_locked", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(conn, "apartments", "opening_balance_locked", "INTEGER NOT NULL DEFAULT 0")
    ensure_column(conn, "apartments", "default_maintenance_amount", "REAL NOT NULL DEFAULT 0")
    ensure_column(conn, "apartments", "opening_balance", "REAL NOT NULL DEFAULT 0")
    migrate_units_table_if_needed(conn)

    legacy_name = cur.execute(
        "SELECT setting_value FROM app_settings WHERE setting_key = 'apartment_name'"
    ).fetchone()
    legacy_owner = cur.execute(
        "SELECT setting_value FROM app_settings WHERE setting_key = 'owner_user_id'"
    ).fetchone()

    if legacy_name and legacy_owner:
        legacy_apartment_name = legacy_name["setting_value"].strip()
        try:
            owner_user_id = int(legacy_owner["setting_value"])
        except (TypeError, ValueError):
            owner_user_id = None

        if legacy_apartment_name and owner_user_id:
            existing_apartment = cur.execute(
                "SELECT apartment_id FROM apartments WHERE owner_user_id = ?",
                (owner_user_id,),
            ).fetchone()
            if existing_apartment:
                apartment_id = existing_apartment["apartment_id"]
                cur.execute(
                    "UPDATE apartments SET apartment_name = ? WHERE apartment_id = ?",
                    (legacy_apartment_name, apartment_id),
                )
            else:
                cursor = cur.execute(
                    "INSERT INTO apartments (apartment_name, owner_user_id) VALUES (?, ?)",
                    (legacy_apartment_name, owner_user_id),
                )
                apartment_id = cursor.lastrowid

            cur.execute(
                "UPDATE users SET apartment_id = ? WHERE user_id = ? AND apartment_id IS NULL",
                (apartment_id, owner_user_id),
            )
            cur.execute(
                "UPDATE units SET apartment_id = ? WHERE apartment_id IS NULL",
                (apartment_id,),
            )
            cur.execute(
                "UPDATE payments SET apartment_id = ? WHERE apartment_id IS NULL",
                (apartment_id,),
            )
            cur.execute(
                "UPDATE expenditures SET apartment_id = ? WHERE apartment_id IS NULL",
                (apartment_id,),
            )

    conn.commit()
    conn.close()


def get_user_apartment_id(user_id: int | None) -> int | None:
    if not user_id:
        return None
    conn = get_conn()
    row = conn.execute(
        "SELECT apartment_id FROM users WHERE user_id = ?",
        (user_id,),
    ).fetchone()
    conn.close()
    if row and row["apartment_id"] is not None:
        return int(row["apartment_id"])
    return None


def get_apartment_name_by_id(apartment_id: int | None) -> str:
    if not apartment_id:
        return ""
    conn = get_conn()
    row = conn.execute(
        "SELECT apartment_name FROM apartments WHERE apartment_id = ?",
        (apartment_id,),
    ).fetchone()
    conn.close()
    if row:
        return row["apartment_name"]
    return ""


def get_apartment_default_maintenance(apartment_id: int | None) -> float:
    if not apartment_id:
        return 0.0
    conn = get_conn()
    row = conn.execute(
        "SELECT default_maintenance_amount FROM apartments WHERE apartment_id = ?",
        (apartment_id,),
    ).fetchone()
    conn.close()
    if row and row["default_maintenance_amount"] is not None:
        return float(row["default_maintenance_amount"])
    return 0.0


def get_apartment_opening_balance(apartment_id: int | None) -> float:
    if not apartment_id:
        return 0.0
    conn = get_conn()
    row = conn.execute(
        "SELECT opening_balance FROM apartments WHERE apartment_id = ?",
        (apartment_id,),
    ).fetchone()
    conn.close()
    if row and row["opening_balance"] is not None:
        return float(row["opening_balance"])
    return 0.0


def set_apartment_opening_balance(apartment_id: int, amount: float) -> None:
    conn = get_conn()
    conn.execute(
        """
        UPDATE apartments
        SET opening_balance = ?, opening_balance_locked = 1
        WHERE apartment_id = ?
        """,
        (amount, apartment_id),
    )
    conn.commit()
    conn.close()


def is_opening_balance_locked(apartment_id: int | None) -> bool:
    if not apartment_id:
        return False
    conn = get_conn()
    row = conn.execute(
        "SELECT opening_balance_locked FROM apartments WHERE apartment_id = ?",
        (apartment_id,),
    ).fetchone()
    conn.close()
    return bool(row and row["opening_balance_locked"])


def get_apartment_start_period(apartment_id: int | None) -> tuple[int, int]:
    today = date.today()
    if not apartment_id:
        return today.year, today.month

    conn = get_conn()
    row = conn.execute(
        "SELECT created_at FROM apartments WHERE apartment_id = ?",
        (apartment_id,),
    ).fetchone()
    conn.close()

    if not row or not row["created_at"]:
        return today.year, today.month

    created_at = str(row["created_at"])
    try:
        created_year = int(created_at[0:4])
        created_month = int(created_at[5:7])
    except (TypeError, ValueError):
        return today.year, today.month

    if created_month < 1 or created_month > 12:
        return today.year, today.month
    return created_year, created_month


def is_before_start_period(year: int, month: int, start_year: int, start_month: int) -> bool:
    return (year, month) < (start_year, start_month)


def is_apartment_name_locked(apartment_id: int | None) -> bool:
    if not apartment_id:
        return False
    conn = get_conn()
    row = conn.execute(
        "SELECT apartment_name_locked FROM apartments WHERE apartment_id = ?",
        (apartment_id,),
    ).fetchone()
    conn.close()
    return bool(row and row["apartment_name_locked"])


def get_owner_profile(apartment_id: int | None) -> dict:
    if not apartment_id:
        return {"display_name": "", "username": ""}

    conn = get_conn()
    row = conn.execute(
        """
        SELECT u.display_name, u.username
        FROM apartments a
        JOIN users u ON u.user_id = a.owner_user_id
        WHERE a.apartment_id = ?
        """,
        (apartment_id,),
    ).fetchone()
    conn.close()

    if not row:
        return {"display_name": "", "username": ""}
    return {"display_name": row["display_name"], "username": row["username"]}


def username_taken(username: str) -> bool:
    conn = get_conn()
    row = conn.execute("SELECT user_id FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    return row is not None


def username_taken_by_other(username: str, user_id: int) -> bool:
    conn = get_conn()
    row = conn.execute(
        "SELECT user_id FROM users WHERE username = ? AND user_id <> ?",
        (username, user_id),
    ).fetchone()
    conn.close()
    return row is not None


def create_user(username: str, password: str, display_name: str, role: str = "admin") -> int:
    conn = get_conn()
    cursor = conn.execute(
        "INSERT INTO users (username, password, display_name, role) VALUES (?, ?, ?, ?)",
        (username, password, display_name, role),
    )
    conn.commit()
    conn.close()
    return int(cursor.lastrowid)


def create_apartment(apartment_name: str, owner_user_id: int) -> int:
    conn = get_conn()
    cursor = conn.execute(
        "INSERT INTO apartments (apartment_name, owner_user_id) VALUES (?, ?)",
        (apartment_name, owner_user_id),
    )
    apartment_id = int(cursor.lastrowid)
    conn.execute(
        "UPDATE users SET apartment_id = ? WHERE user_id = ?",
        (apartment_id, owner_user_id),
    )
    conn.commit()
    conn.close()
    return apartment_id


def update_apartment_name(apartment_id: int, apartment_name: str) -> None:
    conn = get_conn()
    conn.execute(
        "UPDATE apartments SET apartment_name = ?, apartment_name_locked = 1 WHERE apartment_id = ?",
        (apartment_name, apartment_id),
    )
    conn.commit()
    conn.close()


def set_default_maintenance_for_apartment(apartment_id: int, amount: float) -> None:
    conn = get_conn()
    try:
        conn.execute(
            "UPDATE apartments SET default_maintenance_amount = ? WHERE apartment_id = ?",
            (amount, apartment_id),
        )
        conn.execute(
            "UPDATE units SET monthly_maintenance = ? WHERE apartment_id = ?",
            (amount, apartment_id),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def update_owner_account(user_id: int, phone_username: str, display_name: str, new_password: str | None) -> None:
    conn = get_conn()
    if new_password:
        conn.execute(
            """
            UPDATE users
            SET username = ?, display_name = ?, password = ?
            WHERE user_id = ?
            """,
            (phone_username, display_name, new_password, user_id),
        )
    else:
        conn.execute(
            """
            UPDATE users
            SET username = ?, display_name = ?
            WHERE user_id = ?
            """,
            (phone_username, display_name, user_id),
        )
    conn.commit()
    conn.close()


def apartment_has_units(apartment_id: int | None) -> bool:
    if not apartment_id:
        return False
    conn = get_conn()
    row = conn.execute(
        "SELECT COUNT(*) AS count FROM units WHERE apartment_id = ?",
        (apartment_id,),
    ).fetchone()
    conn.close()
    return bool(row and row["count"] > 0)


def replace_apartment_units(
    apartment_id: int,
    floors: int,
    units_per_floor: int,
    owner_names: list[str],
    tenant_names: list[str],
) -> None:
    conn = get_conn()
    try:
        default_amount_row = conn.execute(
            "SELECT default_maintenance_amount FROM apartments WHERE apartment_id = ?",
            (apartment_id,),
        ).fetchone()
        default_amount = (
            float(default_amount_row["default_maintenance_amount"])
            if default_amount_row and default_amount_row["default_maintenance_amount"] is not None
            else 0.0
        )

        conn.execute("DELETE FROM payments WHERE apartment_id = ?", (apartment_id,))
        conn.execute("DELETE FROM maintenance_bills WHERE apartment_id = ?", (apartment_id,))
        conn.execute("DELETE FROM units WHERE apartment_id = ?", (apartment_id,))

        owner_index = 0
        for floor in range(1, floors + 1):
            for unit in range(1, units_per_floor + 1):
                unit_number = f"{floor}{unit:02d}"
                resident_name = owner_names[owner_index]
                tenant_name = tenant_names[owner_index] if owner_index < len(tenant_names) else ""
                conn.execute(
                    """
                    INSERT INTO units (unit_number, resident_name, tenant_name, monthly_maintenance, apartment_id)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (unit_number, resident_name, tenant_name, default_amount, apartment_id),
                )
                owner_index += 1

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def update_apartment_units_occupancy(
    apartment_id: int,
    unit_updates: list[tuple[str, str, int]],
) -> None:
    conn = get_conn()
    try:
        for owner_name, tenant_name, unit_id in unit_updates:
            conn.execute(
                """
                UPDATE units
                SET resident_name = ?, tenant_name = ?
                WHERE unit_id = ? AND apartment_id = ?
                """,
                (owner_name, tenant_name, unit_id, apartment_id),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def get_units_for_apartment(apartment_id: int) -> list[sqlite3.Row]:
    conn = get_conn()
    rows = conn.execute(
        """
         SELECT unit_id, unit_number, resident_name, tenant_name, monthly_maintenance,
             COALESCE(NULLIF(tenant_name, ''), resident_name) AS occupant_name
        FROM units
        WHERE apartment_id = ?
        ORDER BY unit_number
        """,
        (apartment_id,),
    ).fetchall()
    conn.close()
    return rows


def get_unit_payment_status(apartment_id: int, month: int, year: int) -> list[dict]:
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT u.unit_id, u.unit_number, u.resident_name,
             u.tenant_name,
             COALESCE(NULLIF(u.tenant_name, ''), u.resident_name) AS occupant_name,
                             COALESCE(mb.bill_amount, u.monthly_maintenance) AS bill_amount,
               COALESCE(SUM(p.amount_paid), 0) AS paid_amount
        FROM units u
                LEFT JOIN maintenance_bills mb
                    ON mb.unit_id = u.unit_id
                 AND mb.month = ?
                 AND mb.year = ?
                 AND mb.apartment_id = ?
        LEFT JOIN payments p
          ON p.unit_id = u.unit_id
         AND p.month = ?
         AND p.year = ?
         AND p.apartment_id = ?
        WHERE u.apartment_id = ?
        GROUP BY u.unit_id
        ORDER BY u.unit_number
        """,
                (month, year, apartment_id, month, year, apartment_id, apartment_id),
    ).fetchall()
    conn.close()

    result: list[dict] = []
    for row in rows:
        bill_amount = float(row["bill_amount"])
        paid_amount = float(row["paid_amount"])
        pending = max(0.0, bill_amount - paid_amount)
        status = "Paid" if bill_amount > 0 and pending == 0 else "Not Paid"
        result.append(
            {
                "unit_id": row["unit_id"],
                "unit_number": row["unit_number"],
                "resident_name": row["resident_name"],
                "bill_amount": round(bill_amount, 2),
                "paid_amount": round(paid_amount, 2),
                "tenant_name": row["tenant_name"] or "",
                "occupant_name": row["occupant_name"],
                "pending_amount": round(pending, 2),
                "status": status,
            }
        )
    return result


def set_unit_month_payment_status(
    apartment_id: int,
    unit_id: int,
    month: int,
    year: int,
    bill_amount: float,
    payment_status: str,
    payment_mode: str,
    transaction_ref: str,
) -> None:
    normalized_status = payment_status.strip().lower().replace(" ", "_")
    conn = get_conn()
    try:
        unit = conn.execute(
            "SELECT unit_id FROM units WHERE unit_id = ? AND apartment_id = ?",
            (unit_id, apartment_id),
        ).fetchone()
        if not unit:
            raise ValueError("Invalid unit for this apartment")

        conn.execute(
            """
            INSERT INTO maintenance_bills (unit_id, month, year, bill_amount, apartment_id)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(unit_id, month, year, apartment_id)
            DO UPDATE SET bill_amount = excluded.bill_amount
            """,
            (unit_id, month, year, bill_amount, apartment_id),
        )

        conn.execute(
            "DELETE FROM payments WHERE apartment_id = ? AND unit_id = ? AND month = ? AND year = ?",
            (apartment_id, unit_id, month, year),
        )

        if normalized_status == "paid" and bill_amount > 0:
            conn.execute(
                """
                INSERT INTO payments (
                    unit_id, month, year, amount_paid, payment_date, payment_mode, transaction_ref, apartment_id
                ) VALUES (?, ?, ?, ?, date('now'), ?, ?, ?)
                """,
                (unit_id, month, year, bill_amount, payment_mode, transaction_ref, apartment_id),
            )

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def get_expenditures_for_month(apartment_id: int, month: int, year: int) -> list[sqlite3.Row]:
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT expense_id, expense_date, category, vendor_name, amount, invoice_ref, remarks
        FROM expenditures
        WHERE apartment_id = ? AND strftime('%m', expense_date) = ? AND strftime('%Y', expense_date) = ?
        ORDER BY expense_date DESC, expense_id DESC
        """,
        (apartment_id, f"{month:02d}", str(year)),
    ).fetchall()
    conn.close()
    return rows


def create_expenditure(
    apartment_id: int,
    expense_date: str,
    category: str,
    vendor_name: str,
    amount: float,
    invoice_ref: str,
    remarks: str,
) -> None:
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO expenditures (expense_date, category, vendor_name, amount, invoice_ref, remarks, apartment_id)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (expense_date, category, vendor_name, amount, invoice_ref, remarks, apartment_id),
    )
    conn.commit()
    conn.close()


def update_expenditure(
    apartment_id: int,
    expense_id: int,
    expense_date: str,
    category: str,
    vendor_name: str,
    amount: float,
    invoice_ref: str,
    remarks: str,
) -> bool:
    conn = get_conn()
    cursor = conn.execute(
        """
        UPDATE expenditures
        SET expense_date = ?, category = ?, vendor_name = ?, amount = ?, invoice_ref = ?, remarks = ?
        WHERE expense_id = ? AND apartment_id = ?
        """,
        (expense_date, category, vendor_name, amount, invoice_ref, remarks, expense_id, apartment_id),
    )
    conn.commit()
    conn.close()
    return cursor.rowcount > 0


def delete_expenditure(apartment_id: int, expense_id: int) -> bool:
    conn = get_conn()
    cursor = conn.execute(
        "DELETE FROM expenditures WHERE expense_id = ? AND apartment_id = ?",
        (expense_id, apartment_id),
    )
    conn.commit()
    conn.close()
    return cursor.rowcount > 0


def get_income_for_month(apartment_id: int, month: int, year: int) -> list[sqlite3.Row]:
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT income_id, income_date, category, description, amount, remarks
        FROM income
        WHERE apartment_id = ? AND strftime('%m', income_date) = ? AND strftime('%Y', income_date) = ?
        ORDER BY income_date DESC, income_id DESC
        """,
        (apartment_id, f"{month:02d}", str(year)),
    ).fetchall()
    conn.close()
    return rows


def create_income(
    apartment_id: int,
    income_date: str,
    category: str,
    description: str,
    amount: float,
    remarks: str,
) -> None:
    conn = get_conn()
    conn.execute(
        """
        INSERT INTO income (income_date, category, description, amount, remarks, apartment_id)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (income_date, category, description, amount, remarks, apartment_id),
    )
    conn.commit()
    conn.close()


def update_income(
    apartment_id: int,
    income_id: int,
    income_date: str,
    category: str,
    description: str,
    amount: float,
    remarks: str,
) -> bool:
    conn = get_conn()
    cursor = conn.execute(
        """
        UPDATE income
        SET income_date = ?, category = ?, description = ?, amount = ?, remarks = ?
        WHERE income_id = ? AND apartment_id = ?
        """,
        (income_date, category, description, amount, remarks, income_id, apartment_id),
    )
    conn.commit()
    conn.close()
    return cursor.rowcount > 0


def delete_income(apartment_id: int, income_id: int) -> bool:
    conn = get_conn()
    cursor = conn.execute(
        "DELETE FROM income WHERE income_id = ? AND apartment_id = ?",
        (income_id, apartment_id),
    )
    conn.commit()
    conn.close()
    return cursor.rowcount > 0


def is_apartment_owner(user_id: int | None, apartment_id: int | None) -> bool:
    if not user_id or not apartment_id:
        return False
    conn = get_conn()
    row = conn.execute(
        "SELECT apartment_id FROM apartments WHERE apartment_id = ? AND owner_user_id = ?",
        (apartment_id, user_id),
    ).fetchone()
    conn.close()
    return row is not None


def login_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return view_func(*args, **kwargs)

    return wrapper


def owner_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        apartment_id = get_user_apartment_id(session.get("user_id"))
        if not is_apartment_owner(session.get("user_id"), apartment_id):
            return jsonify({"error": "Only the apartment owner can perform this action"}), 403
        return view_func(*args, **kwargs)

    return wrapper


def validate_user(username: str, password: str) -> sqlite3.Row | None:
    conn = get_conn()
    row = conn.execute(
        """
        SELECT user_id, username, display_name, role, apartment_id
        FROM users
        WHERE username = ? AND password = ?
        """,
        (username, password),
    ).fetchone()
    conn.close()
    return row


def month_summary(month: int, year: int, apartment_id: int) -> dict:
    conn = get_conn()
    opening_balance = get_apartment_opening_balance(apartment_id)
    first_day_of_month = f"{year:04d}-{month:02d}-01"

    total_billed = conn.execute(
        """
        SELECT COALESCE(SUM(COALESCE(mb.bill_amount, u.monthly_maintenance)), 0) AS total
        FROM units u
        LEFT JOIN maintenance_bills mb
          ON mb.unit_id = u.unit_id
         AND mb.month = ?
         AND mb.year = ?
         AND mb.apartment_id = ?
        WHERE u.apartment_id = ?
        """,
        (month, year, apartment_id, apartment_id),
    ).fetchone()["total"]

    total_collected = conn.execute(
        """
        SELECT COALESCE(SUM(amount_paid), 0) AS total
        FROM payments
        WHERE apartment_id = ? AND month = ? AND year = ?
        """,
        (apartment_id, month, year),
    ).fetchone()["total"]

    total_expense = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM expenditures
        WHERE apartment_id = ? AND strftime('%m', expense_date) = ? AND strftime('%Y', expense_date) = ?
        """,
        (apartment_id, f"{month:02d}", str(year)),
    ).fetchone()["total"]

    payments = conn.execute(
        """
         SELECT u.unit_number,
             u.resident_name,
             u.tenant_name,
             COALESCE(NULLIF(u.tenant_name, ''), u.resident_name) AS occupant_name,
                         COALESCE(mb.bill_amount, u.monthly_maintenance) AS bill_amount,
               COALESCE(SUM(p.amount_paid), 0) AS paid_amount
        FROM units u
                LEFT JOIN maintenance_bills mb
                    ON mb.unit_id = u.unit_id
                 AND mb.month = ?
                 AND mb.year = ?
                 AND mb.apartment_id = ?
        LEFT JOIN payments p
          ON p.unit_id = u.unit_id
         AND p.month = ?
         AND p.year = ?
         AND p.apartment_id = ?
        WHERE u.apartment_id = ?
        GROUP BY u.unit_id
        ORDER BY u.unit_number
        """,
                (month, year, apartment_id, month, year, apartment_id, apartment_id),
    ).fetchall()

    expenses = conn.execute(
        """
        SELECT expense_date, category, vendor_name, amount, invoice_ref, remarks
        FROM expenditures
        WHERE apartment_id = ? AND strftime('%m', expense_date) = ? AND strftime('%Y', expense_date) = ?
        ORDER BY expense_date DESC
        """,
        (apartment_id, f"{month:02d}", str(year)),
    ).fetchall()

    other_income = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM income
        WHERE apartment_id = ? AND strftime('%m', income_date) = ? AND strftime('%Y', income_date) = ?
        """,
        (apartment_id, f"{month:02d}", str(year)),
    ).fetchone()["total"]

    previous_collected = conn.execute(
        """
        SELECT COALESCE(SUM(amount_paid), 0) AS total
        FROM payments
        WHERE apartment_id = ?
          AND (year < ? OR (year = ? AND month < ?))
        """,
        (apartment_id, year, year, month),
    ).fetchone()["total"]

    previous_other_income = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM income
        WHERE apartment_id = ? AND date(income_date) < date(?)
        """,
        (apartment_id, first_day_of_month),
    ).fetchone()["total"]

    previous_expense = conn.execute(
        """
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM expenditures
        WHERE apartment_id = ? AND date(expense_date) < date(?)
        """,
        (apartment_id, first_day_of_month),
    ).fetchone()["total"]

    income_records = conn.execute(
        """
        SELECT income_date, category, description, amount, remarks
        FROM income
        WHERE apartment_id = ? AND strftime('%m', income_date) = ? AND strftime('%Y', income_date) = ?
        ORDER BY income_date DESC
        """,
        (apartment_id, f"{month:02d}", str(year)),
    ).fetchall()

    conn.close()

    total_pending = total_billed - total_collected
    total_income = total_collected + other_income
    net_balance = total_income - total_expense
    previous_balance = opening_balance + previous_collected + previous_other_income - previous_expense
    cumulative_balance = previous_balance + net_balance
    payment_rows = []
    for row in payments:
        bill_amount = float(row["bill_amount"])
        paid_amount = float(row["paid_amount"])
        pending_amount = max(0.0, bill_amount - paid_amount)
        payment_rows.append(
            {
                "unit_number": row["unit_number"],
                "resident_name": row["resident_name"],
                "tenant_name": row["tenant_name"] or "",
                "occupant_name": row["occupant_name"],
                "bill_amount": round(bill_amount, 2),
                "paid_amount": round(paid_amount, 2),
                "pending_amount": round(pending_amount, 2),
            }
        )

    return {
        "month": month,
        "year": year,
        "total_billed": round(total_billed, 2),
        "total_collected": round(total_collected, 2),
        "total_pending": round(total_pending, 2),
        "total_expense": round(total_expense, 2),
        "total_income": round(total_income, 2),
        "net_balance": round(net_balance, 2),
        "previous_balance": round(previous_balance, 2),
        "cumulative_balance": round(cumulative_balance, 2),
        "payments": payment_rows,
        "expenses": [dict(r) for r in expenses],
        "income": [dict(r) for r in income_records],
    }


def iter_month_periods(start_year: int, start_month: int, end_year: int, end_month: int) -> list[tuple[int, int]]:
    periods: list[tuple[int, int]] = []
    year = start_year
    month = start_month
    while (year, month) <= (end_year, end_month):
        periods.append((year, month))
        month += 1
        if month > 12:
            month = 1
            year += 1
    return periods


def get_reports_dir() -> Path:
    reports_dir = BASE_DIR / "reports"
    reports_dir.mkdir(exist_ok=True)
    return reports_dir


def write_month_report_sheet(ws, apartment_name: str, year: int, month: int, summary: dict) -> None:
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)

    ws.append(["Apartment", apartment_name])
    ws.append(["Period", f"{month:02d}/{year}"])
    ws.append([])

    summary_rows = [
        ("Total Billed", summary["total_billed"]),
        ("Collected", summary["total_collected"]),
        ("Pending", summary["total_pending"]),
        ("Income", summary["total_income"]),
        ("Expenses", summary["total_expense"]),
        ("Balance", summary["net_balance"]),
        ("Previous Balance", summary["previous_balance"]),
        ("Cumulative Balance", summary["cumulative_balance"]),
    ]
    for label, value in summary_rows:
        ws.append([label, value])

    ws.append([])
    payment_header = ["Unit", "Owner", "Tenant", "Bill", "Paid", "Pending"]
    ws.append(payment_header)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for row in summary["payments"]:
        ws.append(
            [
                row["unit_number"],
                row["resident_name"],
                row["tenant_name"] or "-",
                row["bill_amount"],
                row["paid_amount"],
                row["pending_amount"],
            ]
        )

    ws.append([])
    expense_header = ["Date", "Category", "Vendor", "Amount", "Invoice", "Remarks"]
    ws.append(expense_header)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for row in summary["expenses"]:
        ws.append(
            [
                row.get("expense_date", ""),
                row.get("category", ""),
                row.get("vendor_name", ""),
                row.get("amount", 0),
                row.get("invoice_ref", ""),
                row.get("remarks", ""),
            ]
        )

    ws.append([])
    income_header = ["Date", "Category", "Description", "Amount", "Remarks"]
    ws.append(income_header)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
    for row in summary["income"]:
        ws.append(
            [
                row.get("income_date", ""),
                row.get("category", ""),
                row.get("description", ""),
                row.get("amount", 0),
                row.get("remarks", ""),
            ]
        )

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 42)


def build_single_month_workbook(apartment_id: int, apartment_name: str, year: int, month: int) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    summary = month_summary(month, year, apartment_id)
    write_month_report_sheet(ws, apartment_name, year, month, summary)
    return wb


def build_multi_month_workbook(
    apartment_id: int,
    apartment_name: str,
    periods: list[tuple[int, int]],
) -> Workbook:
    wb = Workbook()
    first = True
    for year, month in periods:
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        ws.title = f"{year}-{month:02d}"
        summary = month_summary(month, year, apartment_id)
        write_month_report_sheet(ws, apartment_name, year, month, summary)
    return wb


# OTP Helper Functions
def generate_otp() -> str:
    """Generate a random 6-digit OTP."""
    return ''.join([str(secrets.randbelow(10)) for _ in range(OTP_LENGTH)])


def send_otp_via_email(email_address: str, otp_code: str) -> bool:
    """Send OTP via SendGrid. Returns True if successful, False otherwise."""
    if not sg_client or not SENDGRID_FROM_EMAIL:
        app.logger.warning(f"SendGrid not configured. OTP not sent to {email_address}")
        return False
    
    try:
        message = Mail(
            from_email=SENDGRID_FROM_EMAIL,
            to_emails=email_address,
            subject="Your Maintenance App OTP",
            plain_text_content=f"""Hello,

Your One-Time Password (OTP) for Maintenance Finance App is:

    {otp_code}

This OTP is valid for {OTP_VALIDITY_MINUTES} minutes only.

If you did not request this OTP, please ignore this email.

Best regards,
Maintenance Finance Team""",
            html_content=f"""<html>
<body style="font-family: Arial, sans-serif; color: #333;">
  <h2>Maintenance Finance App - OTP</h2>
  <p>Hello,</p>
  <p>Your One-Time Password (OTP) is:</p>
  <h1 style="background-color: #f0f0f0; padding: 20px; text-align: center; letter-spacing: 5px; font-weight: bold;">
    {otp_code}
  </h1>
  <p><strong>Valid for {OTP_VALIDITY_MINUTES} minutes only.</strong></p>
  <p>If you did not request this OTP, please ignore this email.</p>
  <hr>
  <p style="color: #666; font-size: 12px;">
    Best regards,<br>
    Maintenance Finance Team
  </p>
</body>
</html>"""
        )
        response = sg_client.send(message)
        app.logger.info(f"OTP sent to {email_address}, status code: {response.status_code}")
        return response.status_code in [200, 201, 202]
    except Exception as e:
        app.logger.error(f"Failed to send OTP to {email_address}: {str(e)}")
        return False


def create_otp_session(email_address: str) -> str | None:
    """Create an OTP session and send OTP. Returns OTP code if successful."""
    otp_code = generate_otp()
    expires_at = (datetime.now() + timedelta(minutes=OTP_VALIDITY_MINUTES)).isoformat()
    
    conn = get_conn()
    try:
        conn.execute(
            """
            INSERT INTO otp_sessions (phone_number, otp_code, expires_at)
            VALUES (?, ?, ?)
            """,
            (email_address, otp_code, expires_at)
        )
        conn.commit()
        conn.close()
        
        if send_otp_via_email(email_address, otp_code):
            return otp_code
        else:
            return None
    except Exception as e:
        conn.close()
        app.logger.error(f"Failed to create OTP session: {str(e)}")
        return None


def verify_otp(email_address: str, otp_code: str) -> bool:
    """Verify OTP. Returns True if valid and not expired."""
    conn = get_conn()
    
    otp_row = conn.execute(
        """
        SELECT otp_id, otp_code, expires_at, attempts
        FROM otp_sessions
        WHERE phone_number = ?
        ORDER BY created_at DESC
        LIMIT 1
        """,
        (email_address,)
    ).fetchone()
    
    if not otp_row:
        conn.close()
        return False
    
    # Check expiration
    expires_at = datetime.fromisoformat(otp_row["expires_at"])
    if datetime.now() > expires_at:
        conn.close()
        return False
    
    # Check max attempts
    if otp_row["attempts"] >= 3:
        conn.close()
        return False
    
    # Check OTP code
    if otp_row["otp_code"] != otp_code:
        conn.execute(
            "UPDATE otp_sessions SET attempts = attempts + 1 WHERE otp_id = ?",
            (otp_row["otp_id"],)
        )
        conn.commit()
        conn.close()
        return False
    
    # Mark as verified
    conn.execute(
        "UPDATE otp_sessions SET is_verified = 1 WHERE otp_id = ?",
        (otp_row["otp_id"],)
    )
    conn.commit()
    conn.close()
    return True


def get_user_by_email(email_address: str) -> dict | None:
    """Get user by email address (username)."""
    conn = get_conn()
    user = conn.execute(
        "SELECT user_id, username, display_name, role, apartment_id FROM users WHERE username = ?",
        (email_address,)
    ).fetchone()
    conn.close()
    return dict(user) if user else None


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        step = request.form.get("step", "email")
        
        if step == "email":
            # Step 1: User enters email address
            email_address = request.form.get("email_address", "").strip().lower()
            
            if "@" not in email_address or "." not in email_address:
                return render_template(
                    "login.html",
                    error="Enter a valid email address",
                    step="email"
                )
            
            # Check if user exists
            user = get_user_by_email(email_address)
            if not user:
                return render_template(
                    "login.html",
                    error="Email not found. Please create an account first.",
                    step="email"
                )
            
            # Generate and send OTP
            otp_code = create_otp_session(email_address)
            if not otp_code:
                return render_template(
                    "login.html",
                    error="Failed to send OTP. Please try again.",
                    step="email"
                )
            
            app.logger.info(f"OTP requested: email={email_address}")
            session["login_email"] = email_address
            
            return render_template(
                "login.html",
                step="otp",
                email_address=email_address,
                message=f"OTP sent to {email_address}. Valid for {OTP_VALIDITY_MINUTES} minutes."
            )
        
        elif step == "otp":
            # Step 2: User enters OTP
            email_address = session.get("login_email", "").strip()
            otp_code = request.form.get("otp_code", "").strip()
            
            if not email_address:
                return render_template(
                    "login.html",
                    error="Session expired. Please start again.",
                    step="email"
                )
            
            if len(otp_code) != OTP_LENGTH:
                return render_template(
                    "login.html",
                    step="otp",
                    email_address=email_address,
                    error=f"OTP must be {OTP_LENGTH} digits"
                )
            
            # Verify OTP
            if not verify_otp(email_address, otp_code):
                return render_template(
                    "login.html",
                    step="otp",
                    email_address=email_address,
                    error="Invalid or expired OTP. Please try again."
                )
            
            # OTP verified - log the user in
            user = get_user_by_email(email_address)
            if user:
                session.clear()
                session["user_id"] = int(user["user_id"])
                session["username"] = user["username"]
                session["display_name"] = user["display_name"]
                session["role"] = user["role"]
                session["apartment_id"] = user["apartment_id"]
                
                app.logger.info(f"LOGIN SUCCESS via OTP: user_id={user['user_id']}, email={email_address}")
                
                if user["apartment_id"] is None:
                    return redirect(url_for("setup"))
                if not apartment_has_units(int(user["apartment_id"])):
                    return redirect(url_for("setup_units"))
                return redirect(url_for("dashboard"))
            else:
                return render_template(
                    "login.html",
                    error="User account not found",
                    step="email"
                )
    
    # GET request or initial load
    if "user_id" in session and get_user_apartment_id(session.get("user_id")) is not None:
        apartment_id = get_user_apartment_id(session.get("user_id"))
        if apartment_has_units(apartment_id):
            return redirect(url_for("dashboard"))
        return redirect(url_for("setup_units"))
    
    return render_template("login.html", step="email", error=None)


@app.route("/logout")
def logout():
    app.logger.info(f"LOGOUT: user={session.get('username')}")
    session.clear()
    return redirect(url_for("launch"))


@app.route("/")
def launch():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    return render_template(
        "launch.html",
        apartment_name=get_apartment_name_by_id(apartment_id),
        is_configured=apartment_id is not None,
    )


@app.route("/setup", methods=["GET", "POST"])
def setup():
    current_user_id = session.get("user_id")
    apartment_id = get_user_apartment_id(current_user_id)
    configured = apartment_id is not None
    can_edit_existing = is_apartment_owner(current_user_id, apartment_id)
    owner_profile = get_owner_profile(apartment_id)
    apartment_name = get_apartment_name_by_id(apartment_id)
    apartment_name_locked = is_apartment_name_locked(apartment_id)

    if request.method == "GET":
        return render_template(
            "setup.html",
            error=None,
            apartment_name=apartment_name,
            is_locked=(configured and not can_edit_existing),
            is_configured=configured,
            can_edit_existing=can_edit_existing,
            apartment_name_locked=apartment_name_locked,
            owner_profile=owner_profile,
        )

    if configured and not can_edit_existing:
        return render_template(
            "setup.html",
            error="Apartment details are already configured. Please login with owner account to edit.",
            apartment_name=apartment_name,
            is_locked=True,
            is_configured=True,
            can_edit_existing=False,
            apartment_name_locked=apartment_name_locked,
            owner_profile=owner_profile,
        )

    submitted_apartment_name = request.form.get("apartment_name", "").strip()
    admin_name = request.form.get("admin_name", "").strip()
    admin_phone = request.form.get("admin_phone", "").strip()
    admin_password = request.form.get("admin_password", "").strip()
    confirm_password = request.form.get("confirm_password", "").strip()
    normalized_phone = "".join(ch for ch in admin_phone if ch.isdigit())

    if configured and can_edit_existing:
        if apartment_name_locked:
            return render_template(
                "setup.html",
                error="Apartment name was already updated once and can no longer be changed.",
                apartment_name=apartment_name,
                is_locked=True,
                is_configured=True,
                can_edit_existing=False,
                apartment_name_locked=True,
                owner_profile=owner_profile,
            )

        if not submitted_apartment_name or not admin_name or not normalized_phone:
            return render_template(
                "setup.html",
                error="Apartment name, your name, and mobile number are required.",
                apartment_name=apartment_name,
                is_locked=False,
                is_configured=True,
                can_edit_existing=True,
                apartment_name_locked=False,
                owner_profile=owner_profile,
            )

        if len(normalized_phone) < 10:
            return render_template(
                "setup.html",
                error="Enter a valid mobile number.",
                apartment_name=apartment_name,
                is_locked=False,
                is_configured=True,
                can_edit_existing=True,
                apartment_name_locked=False,
                owner_profile=owner_profile,
            )

        if username_taken_by_other(normalized_phone, int(current_user_id)):
            return render_template(
                "setup.html",
                error="This mobile number is already used by another account.",
                apartment_name=apartment_name,
                is_locked=False,
                is_configured=True,
                can_edit_existing=True,
                apartment_name_locked=False,
                owner_profile=owner_profile,
            )

        new_password = None
        if admin_password or confirm_password:
            if admin_password != confirm_password:
                return render_template(
                    "setup.html",
                    error="Password and confirm password must match.",
                    apartment_name=apartment_name,
                    is_locked=False,
                    is_configured=True,
                    can_edit_existing=True,
                    apartment_name_locked=False,
                    owner_profile=owner_profile,
                )
            new_password = admin_password

        update_apartment_name(int(apartment_id), submitted_apartment_name)
        update_owner_account(int(current_user_id), normalized_phone, admin_name, new_password)
        session["username"] = normalized_phone
        session["display_name"] = admin_name
        return redirect(url_for("dashboard"))

    if not submitted_apartment_name or not admin_name or not admin_phone or not admin_password:
        return render_template(
            "setup.html",
            error="All fields are required.",
            apartment_name="",
            is_locked=False,
            is_configured=False,
            can_edit_existing=False,
            apartment_name_locked=False,
            owner_profile={"display_name": admin_name, "username": admin_phone},
        )

    if admin_password != confirm_password:
        return render_template(
            "setup.html",
            error="Password and confirm password must match.",
            apartment_name="",
            is_locked=False,
            is_configured=False,
            can_edit_existing=False,
            apartment_name_locked=False,
            owner_profile={"display_name": admin_name, "username": admin_phone},
        )

    if len(normalized_phone) < 10:
        return render_template(
            "setup.html",
            error="Enter a valid mobile number.",
            apartment_name="",
            is_locked=False,
            is_configured=False,
            can_edit_existing=False,
            apartment_name_locked=False,
            owner_profile={"display_name": admin_name, "username": admin_phone},
        )

    if username_taken(normalized_phone):
        return render_template(
            "setup.html",
            error="This mobile number already has an apartment account. Please login.",
            apartment_name="",
            is_locked=False,
            is_configured=False,
            can_edit_existing=False,
            apartment_name_locked=False,
            owner_profile={"display_name": admin_name, "username": admin_phone},
        )

    user_id = create_user(normalized_phone, admin_password, admin_name)
    apartment_id = create_apartment(submitted_apartment_name, user_id)
    session.clear()
    session["user_id"] = user_id
    session["username"] = normalized_phone
    session["display_name"] = admin_name
    session["role"] = "admin"
    session["apartment_id"] = apartment_id
    return redirect(url_for("setup_units"))


@app.route("/setup/units", methods=["GET", "POST"])
@login_required
def setup_units():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    if not is_apartment_owner(session.get("user_id"), apartment_id):
        return redirect(url_for("dashboard"))

    if request.method == "GET":
        return render_template(
            "setup_units.html",
            apartment_name=get_apartment_name_by_id(apartment_id),
            error=None,
            step="counts",
            floors="",
            units_per_floor="",
            unit_labels=[],
            owner_names=[],
            tenant_names=[],
        )

    step = request.form.get("step", "counts")
    floors_raw = request.form.get("floors", "").strip()
    units_raw = request.form.get("units_per_floor", "").strip()

    try:
        floors = int(floors_raw)
        units_per_floor = int(units_raw)
    except ValueError:
        floors = 0
        units_per_floor = 0

    if floors <= 0 or units_per_floor <= 0:
        return render_template(
            "setup_units.html",
            apartment_name=get_apartment_name_by_id(apartment_id),
            error="Enter valid numbers for floors and units per floor.",
            step="counts",
            floors=floors_raw,
            units_per_floor=units_raw,
            unit_labels=[],
            owner_names=[],
            tenant_names=[],
        )

    unit_labels = [f"{floor}{unit:02d}" for floor in range(1, floors + 1) for unit in range(1, units_per_floor + 1)]

    if step == "counts":
        return render_template(
            "setup_units.html",
            apartment_name=get_apartment_name_by_id(apartment_id),
            error=None,
            step="owners",
            floors=floors,
            units_per_floor=units_per_floor,
            unit_labels=unit_labels,
            owner_names=["" for _ in unit_labels],
            tenant_names=["" for _ in unit_labels],
        )

    owner_names = [request.form.get(f"owner_name_{index}", "").strip() for index in range(len(unit_labels))]
    tenant_names = [request.form.get(f"tenant_name_{index}", "").strip() for index in range(len(unit_labels))]
    if any(not name for name in owner_names):
        return render_template(
            "setup_units.html",
            apartment_name=get_apartment_name_by_id(apartment_id),
            error="Enter owner name for every unit.",
            step="owners",
            floors=floors,
            units_per_floor=units_per_floor,
            unit_labels=unit_labels,
            owner_names=owner_names,
            tenant_names=tenant_names,
        )

    replace_apartment_units(apartment_id, floors, units_per_floor, owner_names, tenant_names)
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
@login_required
def dashboard():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))
    if not apartment_has_units(apartment_id):
        return redirect(url_for("setup_units"))

    today = date.today()
    start_year, start_month = get_apartment_start_period(apartment_id)
    selected_month = request.args.get("month", str(today.month)).strip()
    selected_year = request.args.get("year", str(today.year)).strip()

    try:
        month = int(selected_month)
    except ValueError:
        month = today.month

    try:
        year = int(selected_year)
    except ValueError:
        year = today.year

    if month < 1 or month > 12:
        month = today.month
    if year < 2000 or year > 2100:
        year = today.year
    if is_before_start_period(year, month, start_year, start_month):
        year = start_year
        month = start_month

    summary = month_summary(month, year, apartment_id)
    is_owner = is_apartment_owner(session.get("user_id"), apartment_id)
    return render_template(
        "index.html",
        summary=summary,
        selected_month=month,
        selected_year=year,
        start_month=start_month,
        start_year=start_year,
        apartment_name=get_apartment_name_by_id(apartment_id),
        opening_balance=get_apartment_opening_balance(apartment_id),
        user_display_name=session.get("display_name", "User"),
        can_edit_settings=is_owner and not is_apartment_name_locked(apartment_id),
        can_manage_payments=is_owner,
        can_manage_units=is_owner,
        can_manage_expenditures=is_owner,
        can_view_reports=True,
        can_set_opening_balance=is_owner and not is_opening_balance_locked(apartment_id),
        opening_balance_locked=is_opening_balance_locked(apartment_id),
    )


@app.route("/reports")
@login_required
def reports_page():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    apartment_name = get_apartment_name_by_id(apartment_id)
    start_year, start_month = get_apartment_start_period(apartment_id)
    today = date.today()

    reports_dir = get_reports_dir()
    prefix = f"report_{apartment_id}_"
    files = []
    for file in sorted(reports_dir.glob(f"{prefix}*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        stat = file.stat()
        files.append(
            {
                "name": file.name,
                "size_kb": round(stat.st_size / 1024, 2),
                "updated": date.fromtimestamp(stat.st_mtime).isoformat(),
            }
        )

    return render_template(
        "reports.html",
        apartment_name=apartment_name,
        start_year=start_year,
        start_month=start_month,
        current_year=today.year,
        current_month=today.month,
        files=files,
        message=request.args.get("message", ""),
        error=request.args.get("error", ""),
    )


@app.route("/reports/generate", methods=["POST"])
@login_required
def generate_reports():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    apartment_name = get_apartment_name_by_id(apartment_id)
    start_year, start_month = get_apartment_start_period(apartment_id)
    today = date.today()
    reports_dir = get_reports_dir()
    action = request.form.get("action", "")

    try:
        if action == "single_month":
            month = int(request.form.get("month", str(today.month)))
            year = int(request.form.get("year", str(today.year)))
            if month < 1 or month > 12:
                raise ValueError("Month must be between 1 and 12")
            if year < 2000 or year > 2100:
                raise ValueError("Enter a valid year")
            if is_before_start_period(year, month, start_year, start_month):
                raise ValueError(f"History starts from {start_month:02d}/{start_year}")

            wb = build_single_month_workbook(apartment_id, apartment_name, year, month)
            out_file = reports_dir / f"report_{apartment_id}_{year}_{month:02d}.xlsx"
            wb.save(out_file)
            return redirect(url_for("reports_page", message=f"Generated {out_file.name}"))

        if action == "separate_all_months":
            periods = iter_month_periods(start_year, start_month, today.year, today.month)
            count = 0
            for year, month in periods:
                wb = build_single_month_workbook(apartment_id, apartment_name, year, month)
                out_file = reports_dir / f"report_{apartment_id}_{year}_{month:02d}.xlsx"
                wb.save(out_file)
                count += 1
            return redirect(url_for("reports_page", message=f"Generated {count} monthly report files"))

        if action == "all_months_workbook":
            periods = iter_month_periods(start_year, start_month, today.year, today.month)
            wb = build_multi_month_workbook(apartment_id, apartment_name, periods)
            out_file = reports_dir / f"report_{apartment_id}_all_months_{today.year}{today.month:02d}.xlsx"
            wb.save(out_file)
            return redirect(url_for("reports_page", message=f"Generated {out_file.name}"))

        if action == "generate_all_zip":
            periods = iter_month_periods(start_year, start_month, today.year, today.month)
            monthly_files: list[Path] = []
            for year, month in periods:
                wb = build_single_month_workbook(apartment_id, apartment_name, year, month)
                out_file = reports_dir / f"report_{apartment_id}_{year}_{month:02d}.xlsx"
                wb.save(out_file)
                monthly_files.append(out_file)

            zip_name = f"report_{apartment_id}_all_months_bundle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            zip_path = reports_dir / zip_name
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as bundle:
                for file_path in monthly_files:
                    bundle.write(file_path, arcname=file_path.name)

            return send_from_directory(reports_dir, zip_name, as_attachment=True)

        if action == "custom_range":
            from_month = int(request.form.get("from_month", str(start_month)))
            from_year = int(request.form.get("from_year", str(start_year)))
            to_month = int(request.form.get("to_month", str(today.month)))
            to_year = int(request.form.get("to_year", str(today.year)))

            if from_month < 1 or from_month > 12 or to_month < 1 or to_month > 12:
                raise ValueError("Month must be between 1 and 12")
            if from_year < 2000 or to_year < 2000:
                raise ValueError("Enter a valid year")
            if is_before_start_period(from_year, from_month, start_year, start_month):
                raise ValueError(f"Range must start from {start_month:02d}/{start_year} or later")
            if (from_year, from_month) > (to_year, to_month):
                raise ValueError("From period must be before or equal to To period")
            if (to_year, to_month) > (today.year, today.month):
                raise ValueError("To period cannot be in the future")

            periods = iter_month_periods(from_year, from_month, to_year, to_month)
            wb = build_multi_month_workbook(apartment_id, apartment_name, periods)
            out_file = reports_dir / f"report_{apartment_id}_{from_year}{from_month:02d}_to_{to_year}{to_month:02d}.xlsx"
            wb.save(out_file)
            return redirect(url_for("reports_page", message=f"Generated {out_file.name}"))

        return redirect(url_for("reports_page", error="Invalid report action"))
    except ValueError as exc:
        return redirect(url_for("reports_page", error=str(exc)))
    except Exception:
        return redirect(url_for("reports_page", error="Unable to generate report. Please try again."))


@app.route("/reports/download/<path:filename>")
@login_required
def download_report(filename: str):
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    safe_name = Path(filename).name
    if safe_name != filename:
        return jsonify({"error": "Invalid file name"}), 400

    if not safe_name.startswith(f"report_{apartment_id}_"):
        return jsonify({"error": "File not found"}), 404

    reports_dir = get_reports_dir()
    file_path = reports_dir / safe_name
    if not file_path.exists():
        return jsonify({"error": "File not found"}), 404

    return send_from_directory(reports_dir, safe_name, as_attachment=True)


@app.route("/units/manage", methods=["GET", "POST"])
@owner_required
def manage_units():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    message = ""
    error = ""

    if request.method == "POST":
        try:
            units = get_units_for_apartment(apartment_id)
            updates: list[tuple[str, str, int]] = []

            for unit in units:
                unit_id = int(unit["unit_id"])
                unit_number = unit["unit_number"]
                owner_name = request.form.get(f"owner_name_{unit_id}", "").strip()
                tenant_name = request.form.get(f"tenant_name_{unit_id}", "").strip()

                if not owner_name:
                    raise ValueError(f"Owner name is required for unit {unit_number}")

                updates.append((owner_name, tenant_name, unit_id))

            update_apartment_units_occupancy(apartment_id, updates)
            message = "Unit owner and tenant details updated successfully."
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = "Unable to update unit details. Please try again."

    units = get_units_for_apartment(apartment_id)
    return render_template(
        "units_manage.html",
        apartment_name=get_apartment_name_by_id(apartment_id),
        units=units,
        message=message,
        error=error,
    )


@app.route("/payments/manage", methods=["GET", "POST"])
@owner_required
def manage_payments():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    today = date.today()
    month = int(request.values.get("month", today.month))
    year = int(request.values.get("year", today.year))
    start_year, start_month = get_apartment_start_period(apartment_id)
    message = request.values.get("message", "")
    error = ""
    default_maintenance_amount = get_apartment_default_maintenance(apartment_id)

    if is_before_start_period(year, month, start_year, start_month):
        year = start_year
        month = start_month
        error = f"History starts from {start_month:02d}/{start_year}."

    if request.method == "POST":
        action = request.form.get("action", "bulk_update")
        try:
            if action == "set_default":
                default_amount = float(request.form.get("default_maintenance_amount", "0"))
                if default_amount < 0:
                    raise ValueError("Default maintenance amount cannot be negative")
                set_default_maintenance_for_apartment(apartment_id, default_amount)
                default_maintenance_amount = default_amount
                message = "Default maintenance amount saved and applied to all units."
            else:
                selected_units_raw = request.form.getlist("selected_units")
                bill_amount = float(request.form.get("bill_amount", "0"))
                payment_status = request.form.get("bulk_status", "not_paid")
                apply_action = request.form.get("apply_action", "")
                payment_mode = request.form.get("payment_mode", "Manual")
                transaction_ref = request.form.get("transaction_ref", "").strip()

                normalized_status = payment_status.strip().lower().replace(" ", "_")
                if apply_action == "mark_paid":
                    normalized_status = "paid"
                elif apply_action == "mark_unpaid":
                    normalized_status = "not_paid"

                if month < 1 or month > 12:
                    raise ValueError("Month must be between 1 and 12")
                if year < 2000 or year > 2100:
                    raise ValueError("Enter a valid year")
                if is_before_start_period(year, month, start_year, start_month):
                    raise ValueError(f"History starts from {start_month:02d}/{start_year}")
                if bill_amount < 0:
                    raise ValueError("Bill amount cannot be negative")
                if normalized_status not in ["paid", "not_paid"]:
                    raise ValueError("Invalid payment status")
                if not selected_units_raw:
                    raise ValueError("Select at least one unit.")

                selected_units: list[int] = []
                for value in selected_units_raw:
                    try:
                        selected_units.append(int(value))
                    except ValueError:
                        continue

                if not selected_units:
                    raise ValueError("Select valid units.")

                updated_count = 0
                for unit_id in selected_units:
                    set_unit_month_payment_status(
                        apartment_id=apartment_id,
                        unit_id=unit_id,
                        month=month,
                        year=year,
                        bill_amount=bill_amount,
                        payment_status=normalized_status,
                        payment_mode=payment_mode,
                        transaction_ref=transaction_ref,
                    )
                    updated_count += 1

                app.logger.info(f"PAYMENT UPDATE: user={session.get('username')}, apartment_id={apartment_id}, units={len(selected_units)}, status={normalized_status}, amount={bill_amount}, month={month}/{year}")
                message = f"Updated {updated_count} unit(s) successfully."
        except ValueError as exc:
            error = str(exc)
            app.logger.warning(f"PAYMENT UPDATE ERROR: user={session.get('username')}, error={error}")
        except Exception as e:
            error = "Unable to save maintenance status. Please try again."
            app.logger.error(f"PAYMENT UPDATE EXCEPTION: user={session.get('username')}, error={str(e)}")

    units = get_units_for_apartment(apartment_id)
    status_rows = get_unit_payment_status(apartment_id, month, year)

    return render_template(
        "payments_manage.html",
        apartment_name=get_apartment_name_by_id(apartment_id),
        month=month,
        year=year,
        units=units,
        status_rows=status_rows,
        message=message,
        error=error,
        default_maintenance_amount=default_maintenance_amount,
        start_month=start_month,
        start_year=start_year,
    )


@app.route("/expenditures/manage", methods=["GET", "POST"])
@owner_required
def manage_expenditures():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    today = date.today()
    month = int(request.values.get("month", today.month))
    year = int(request.values.get("year", today.year))
    start_year, start_month = get_apartment_start_period(apartment_id)
    message = request.values.get("message", "")
    error = ""

    if is_before_start_period(year, month, start_year, start_month):
        year = start_year
        month = start_month
        error = f"History starts from {start_month:02d}/{start_year}."

    if request.method == "POST":
        action = request.form.get("action", "create")
        try:
            if month < 1 or month > 12:
                raise ValueError("Month must be between 1 and 12")
            if year < 2000 or year > 2100:
                raise ValueError("Enter a valid year")
            if is_before_start_period(year, month, start_year, start_month):
                raise ValueError(f"History starts from {start_month:02d}/{start_year}")

            if action == "delete":
                expense_id = int(request.form.get("expense_id", "0"))
                if expense_id <= 0:
                    raise ValueError("Invalid expenditure id")
                if not delete_expenditure(apartment_id, expense_id):
                    raise ValueError("Expenditure not found")
                message = "Expenditure deleted successfully."
            else:
                expense_date = request.form.get("expense_date", "").strip()
                category = request.form.get("category", "").strip()
                vendor_name = request.form.get("vendor_name", "").strip()
                amount = float(request.form.get("amount", "0"))
                invoice_ref = request.form.get("invoice_ref", "").strip()
                remarks = request.form.get("remarks", "").strip()

                if not expense_date or not category or not vendor_name:
                    raise ValueError("Date, category, and vendor are required")
                if amount < 0:
                    raise ValueError("Amount cannot be negative")

                if action == "update":
                    expense_id = int(request.form.get("expense_id", "0"))
                    if expense_id <= 0:
                        raise ValueError("Invalid expenditure id")
                    if not update_expenditure(
                        apartment_id,
                        expense_id,
                        expense_date,
                        category,
                        vendor_name,
                        amount,
                        invoice_ref,
                        remarks,
                    ):
                        raise ValueError("Expenditure not found")
                    app.logger.info(f"EXPENSE UPDATE: user={session.get('username')}, apartment_id={apartment_id}, expense_id={expense_id}, category={category}, amount={amount}")
                    message = "Expenditure updated successfully."
                else:
                    create_expenditure(
                        apartment_id,
                        expense_date,
                        category,
                        vendor_name,
                        amount,
                        invoice_ref,
                        remarks,
                    )
                    app.logger.info(f"EXPENSE CREATED: user={session.get('username')}, apartment_id={apartment_id}, category={category}, amount={amount}")
                    message = "Expenditure added successfully."
        except ValueError as exc:
            error = str(exc)
            app.logger.warning(f"EXPENSE ERROR: user={session.get('username')}, error={error}")
        except Exception as e:
            error = "Unable to save expenditure. Please try again."
            app.logger.error(f"EXPENSE EXCEPTION: user={session.get('username')}, error={str(e)}")

    rows = get_expenditures_for_month(apartment_id, month, year)

    return render_template(
        "expenditures_manage.html",
        apartment_name=get_apartment_name_by_id(apartment_id),
        month=month,
        year=year,
        rows=rows,
        message=message,
        error=error,
        start_month=start_month,
        start_year=start_year,
    )


@app.route("/income/manage", methods=["GET", "POST"])
@owner_required
def manage_income():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return redirect(url_for("setup"))

    today = date.today()
    month = int(request.values.get("month", today.month))
    year = int(request.values.get("year", today.year))
    start_year, start_month = get_apartment_start_period(apartment_id)
    message = request.values.get("message", "")
    error = ""

    if is_before_start_period(year, month, start_year, start_month):
        year = start_year
        month = start_month
        error = f"History starts from {start_month:02d}/{start_year}."

    if request.method == "POST":
        action = request.form.get("action", "create")
        try:
            if month < 1 or month > 12:
                raise ValueError("Month must be between 1 and 12")
            if year < 2000 or year > 2100:
                raise ValueError("Enter a valid year")
            if is_before_start_period(year, month, start_year, start_month):
                raise ValueError(f"History starts from {start_month:02d}/{start_year}")

            if action == "delete":
                income_id = int(request.form.get("income_id", "0"))
                if income_id <= 0:
                    raise ValueError("Invalid income id")
                if not delete_income(apartment_id, income_id):
                    raise ValueError("Income record not found")
                message = "Income deleted successfully."
            else:
                income_date = request.form.get("income_date", "").strip()
                category = request.form.get("category", "").strip()
                description = request.form.get("description", "").strip()
                amount = float(request.form.get("amount", "0"))
                remarks = request.form.get("remarks", "").strip()

                if not income_date or not category or not description:
                    raise ValueError("Date, category, and description are required")
                if amount < 0:
                    raise ValueError("Amount cannot be negative")

                if action == "update":
                    income_id = int(request.form.get("income_id", "0"))
                    if income_id <= 0:
                        raise ValueError("Invalid income id")
                    if not update_income(
                        apartment_id,
                        income_id,
                        income_date,
                        category,
                        description,
                        amount,
                        remarks,
                    ):
                        raise ValueError("Income record not found")
                    message = "Income updated successfully."
                else:
                    create_income(
                        apartment_id,
                        income_date,
                        category,
                        description,
                        amount,
                        remarks,
                    )
                    message = "Income added successfully."
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = "Unable to save income. Please try again."

    rows = get_income_for_month(apartment_id, month, year)

    return render_template(
        "income_manage.html",
        apartment_name=get_apartment_name_by_id(apartment_id),
        month=month,
        year=year,
        rows=rows,
        message=message,
        error=error,
        start_month=start_month,
        start_year=start_year,
    )


@app.route("/settings/apartment-name", methods=["POST"])
@owner_required
def update_apartment_name_route():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    apartment_name = request.form.get("apartment_name", "").strip()
    if apartment_id and apartment_name and not is_apartment_name_locked(apartment_id):
        update_apartment_name(apartment_id, apartment_name)
    return redirect(url_for("dashboard"))


@app.route("/settings/opening-balance", methods=["POST"])
@owner_required
def update_opening_balance_route():
    apartment_id = get_user_apartment_id(session.get("user_id"))
    try:
        opening_balance = float(request.form.get("opening_balance", "0"))
    except ValueError:
        opening_balance = 0.0

    if apartment_id and opening_balance >= 0 and not is_opening_balance_locked(apartment_id):
        set_apartment_opening_balance(apartment_id, opening_balance)

    return redirect(url_for("dashboard"))


@app.route("/api/summary/<int:year>/<int:month>")
@login_required
def api_summary(year: int, month: int):
    if month < 1 or month > 12:
        return jsonify({"error": "Month must be 1-12"}), 400

    apartment_id = get_user_apartment_id(session.get("user_id"))
    if apartment_id is None:
        return jsonify({"error": "No apartment configured for this account"}), 400

    start_year, start_month = get_apartment_start_period(apartment_id)
    if is_before_start_period(year, month, start_year, start_month):
        return (
            jsonify({"error": f"History starts from {start_month:02d}/{start_year}"}),
            400,
        )

    payload = month_summary(month, year, apartment_id)
    payload["apartment_name"] = get_apartment_name_by_id(apartment_id)
    return jsonify(payload)


@app.route("/admin/logs")
@login_required
def view_logs():
    """View application logs (admin only)."""
    try:
        # Only owner can view logs
        if not session.get("apartment_id"):
            return jsonify({"error": "Unauthorized"}), 403
        
        if not LOG_FILE.exists():
            return jsonify({"logs": [], "message": "No logs available yet"})
        
        # Read last 100 lines of the log file
        with open(LOG_FILE, 'r') as f:
            lines = f.readlines()
        
        recent_logs = lines[-100:] if len(lines) > 100 else lines
        
        app.logger.info(f"LOGS VIEWED: user={session.get('username')}")
        
        return jsonify({
            "logs": recent_logs,
            "total_lines": len(lines),
            "message": f"Showing last {len(recent_logs)} lines"
        })
    except Exception as e:
        app.logger.error(f"Error viewing logs: {str(e)}")
        return jsonify({"error": str(e)}), 500


try:
    init_db()
except Exception as e:
    app.logger.error(f"Database initialization failed during startup: {str(e)}")
    raise


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
